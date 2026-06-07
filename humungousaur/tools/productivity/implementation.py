from __future__ import annotations

from datetime import datetime, timezone
import json
from pathlib import Path
import re
from typing import Any
from uuid import uuid4

from humungousaur.config import AgentConfig
from humungousaur.schemas import ActionStatus, RiskLevel, ToolResult
from humungousaur.tools.base import Tool, object_input_schema


EMAIL_ADDRESS_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")
EMAIL_DRAFT_MAX_RECIPIENTS = 50
EMAIL_DRAFT_MAX_BODY_CHARS = 50_000
XLSX_MAX_ROWS = 5_000
XLSX_MAX_COLUMNS = 100
API_OPERATION_MAX_RECORDS = 100
API_OPERATION_MAX_BLOCKS = 500
API_OPERATION_BODY_CHARS = 120_000


class EmailDraftPrepareTool(Tool):
    def __init__(self) -> None:
        super().__init__(
            name="email_draft_prepare",
            description=(
                "Prepare a local email draft artifact from explicit recipients, subject, and body. "
                "This does not send email; use it for approval-ready Gmail, Outlook, IMAP, or SMTP drafts."
            ),
            risk_level=RiskLevel.MEDIUM,
            input_schema=_email_draft_schema(),
            capability_group="productivity",
        )

    def execute(self, tool_input: dict[str, Any], config: AgentConfig) -> ToolResult:
        return _prepare_email_draft(tool_input, config, provider="email")


class GmailDraftPrepareTool(Tool):
    def __init__(self) -> None:
        super().__init__(
            name="gmail_draft_prepare",
            description=(
                "Prepare a Gmail-ready local draft artifact from explicit recipients, subject, and body. "
                "This validates the draft and records the next gog/native Gmail command shape, but never sends."
            ),
            risk_level=RiskLevel.MEDIUM,
            input_schema=_email_draft_schema(),
            capability_group="productivity",
        )

    def execute(self, tool_input: dict[str, Any], config: AgentConfig) -> ToolResult:
        return _prepare_email_draft(tool_input, config, provider="gmail")


class XlsxWorkbookCreateTool(Tool):
    def __init__(self) -> None:
        super().__init__(
            name="xlsx_workbook_create",
            description=(
                "Create a native XLSX workbook artifact with sheets, rows, formulas, and optional number formats. "
                "Use for Excel-style workbook tasks before sharing or editing source files."
            ),
            risk_level=RiskLevel.MEDIUM,
            input_schema=object_input_schema(
                {
                    "filename": {"type": "string", "description": "Output filename under data_dir/spreadsheets."},
                    "sheets": {
                        "type": "array",
                        "items": {"type": "object"},
                        "description": "Sheet specs: name, rows, formulas, freeze_panes, widths.",
                    },
                    "reason": {"type": "string", "description": "Why this workbook should be created."},
                },
                required=["sheets", "reason"],
            ),
            capability_group="productivity",
        )

    def execute(self, tool_input: dict[str, Any], config: AgentConfig) -> ToolResult:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font
        except Exception as exc:
            return ToolResult(self.name, ActionStatus.FAILED, self.risk_level, "openpyxl is required for XLSX creation.", error=str(exc))

        sheets = tool_input.get("sheets")
        if not isinstance(sheets, list) or not sheets:
            return ToolResult(self.name, ActionStatus.FAILED, self.risk_level, "At least one sheet spec is required.")
        filename = _safe_filename(str(tool_input.get("filename") or "humungousaur-workbook.xlsx"), ".xlsx")
        path = (config.normalized().data_dir / "spreadsheets" / filename).resolve()
        if not _is_within(path, config.normalized().allowed_write_roots):
            return ToolResult(self.name, ActionStatus.BLOCKED, self.risk_level, "Workbook path is outside allowed write roots.")
        if config.dry_run:
            return ToolResult(self.name, ActionStatus.SKIPPED, self.risk_level, f"Dry run: would create workbook {path}.", {"path": str(path), "sheets": sheets})

        workbook = Workbook()
        default_sheet = workbook.active
        workbook.remove(default_sheet)
        summaries: list[dict[str, Any]] = []
        for index, spec in enumerate(sheets[:20], start=1):
            if not isinstance(spec, dict):
                continue
            sheet_name = _safe_sheet_name(str(spec.get("name") or f"Sheet{index}"), workbook.sheetnames)
            worksheet = workbook.create_sheet(sheet_name)
            rows = spec.get("rows") or []
            if not isinstance(rows, list):
                rows = []
            if len(rows) > XLSX_MAX_ROWS:
                return ToolResult(self.name, ActionStatus.BLOCKED, self.risk_level, f"Sheet {sheet_name} exceeds row limit.")
            for row in rows:
                if not isinstance(row, list):
                    row = [row]
                if len(row) > XLSX_MAX_COLUMNS:
                    return ToolResult(self.name, ActionStatus.BLOCKED, self.risk_level, f"Sheet {sheet_name} exceeds column limit.")
                worksheet.append([_xlsx_cell_value(value) for value in row])
            formulas = spec.get("formulas") or []
            if isinstance(formulas, list):
                for formula in formulas[:500]:
                    if not isinstance(formula, dict):
                        continue
                    cell = str(formula.get("cell") or "").strip()
                    value = str(formula.get("formula") or "").strip()
                    if cell and value:
                        worksheet[cell] = value if value.startswith("=") else f"={value}"
            if worksheet.max_row >= 1:
                for cell in worksheet[1]:
                    cell.font = Font(bold=True)
            freeze_panes = str(spec.get("freeze_panes") or "").strip()
            if freeze_panes:
                worksheet.freeze_panes = freeze_panes
            widths = spec.get("widths") or {}
            if isinstance(widths, dict):
                for column, width in widths.items():
                    try:
                        worksheet.column_dimensions[str(column).upper()].width = max(1, min(float(width), 80))
                    except (TypeError, ValueError):
                        continue
            summaries.append({"name": sheet_name, "rows": worksheet.max_row, "columns": worksheet.max_column})
        if not workbook.sheetnames:
            return ToolResult(self.name, ActionStatus.FAILED, self.risk_level, "No valid sheets were provided.")
        path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(path)
        return ToolResult(
            self.name,
            ActionStatus.SUCCEEDED,
            self.risk_level,
            f"Created XLSX workbook {path}.",
            {"path": str(path), "sheets": summaries, "source": "xlsx_workbook_create"},
        )


class XlsxWorkbookInspectTool(Tool):
    def __init__(self) -> None:
        super().__init__(
            name="xlsx_workbook_inspect",
            description="Inspect a local XLSX workbook artifact for sheets, dimensions, headers, formulas, and sample rows.",
            risk_level=RiskLevel.LOW,
            input_schema=object_input_schema(
                {
                    "path": {"type": "string", "description": "Workspace-relative or allowed absolute XLSX path."},
                    "sample_rows": {"type": "integer", "minimum": 1, "maximum": 20},
                },
                required=["path"],
            ),
            capability_group="productivity",
        )

    def execute(self, tool_input: dict[str, Any], config: AgentConfig) -> ToolResult:
        try:
            from openpyxl import load_workbook
        except Exception as exc:
            return ToolResult(self.name, ActionStatus.FAILED, self.risk_level, "openpyxl is required for XLSX inspection.", error=str(exc))
        path = _resolve_allowed_path(config.normalized(), str(tool_input.get("path") or ""))
        if not _is_within(path, config.normalized().allowed_read_roots + config.normalized().allowed_write_roots):
            return ToolResult(self.name, ActionStatus.BLOCKED, self.risk_level, "Workbook path is outside allowed roots.")
        if not path.exists() or path.suffix.lower() != ".xlsx":
            return ToolResult(self.name, ActionStatus.FAILED, self.risk_level, "Workbook does not exist or is not an XLSX file.")
        sample_rows = max(1, min(int(tool_input.get("sample_rows") or 5), 20))
        try:
            workbook = load_workbook(path, data_only=False, read_only=True)
        except Exception as exc:
            return ToolResult(self.name, ActionStatus.FAILED, self.risk_level, "Workbook could not be opened.", error=str(exc))
        sheets: list[dict[str, Any]] = []
        for worksheet in workbook.worksheets:
            rows = []
            formulas = []
            for row_index, row in enumerate(worksheet.iter_rows(values_only=False), start=1):
                if row_index <= sample_rows:
                    rows.append([cell.value for cell in row])
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        formulas.append({"cell": cell.coordinate, "formula": cell.value})
                if row_index >= sample_rows and len(formulas) >= 50:
                    break
            sheets.append(
                {
                    "name": worksheet.title,
                    "max_row": worksheet.max_row,
                    "max_column": worksheet.max_column,
                    "sample_rows": rows,
                    "formulas": formulas[:50],
                }
            )
        workbook.close()
        return ToolResult(
            self.name,
            ActionStatus.SUCCEEDED,
            self.risk_level,
            f"Inspected XLSX workbook {path}.",
            {"path": str(path), "sheets": sheets, "source": "xlsx_workbook_inspect"},
        )


class NotionOperationPrepareTool(Tool):
    def __init__(self) -> None:
        super().__init__(
            name="notion_operation_prepare",
            description=(
                "Prepare a Notion API operation artifact for page/database workflows with validated target IDs, payload preview, "
                "approval state, and endpoint metadata. This does not call Notion."
            ),
            risk_level=RiskLevel.MEDIUM,
            input_schema=object_input_schema(
                {
                    "operation": {
                        "type": "string",
                        "enum": ["create_page", "update_page", "append_blocks", "query_database", "update_database_schema"],
                    },
                    "page_id": {"type": "string"},
                    "database_id": {"type": "string"},
                    "parent_page_id": {"type": "string"},
                    "title": {"type": "string"},
                    "properties": {"type": "object"},
                    "blocks": {"type": "array", "items": {"type": "object"}},
                    "filter": {"type": "object"},
                    "sorts": {"type": "array", "items": {"type": "object"}},
                    "schema": {"type": "object", "description": "Database property schema changes for update_database_schema."},
                    "notion_version": {"type": "string"},
                    "filename": {"type": "string"},
                    "reason": {"type": "string"},
                },
                required=["operation", "reason"],
            ),
            capability_group="productivity",
        )

    def execute(self, tool_input: dict[str, Any], config: AgentConfig) -> ToolResult:
        normalized = config.normalized()
        operation = str(tool_input.get("operation") or "").strip()
        reason = str(tool_input.get("reason") or "").strip()
        if not operation or not reason:
            return ToolResult(self.name, ActionStatus.FAILED, self.risk_level, "Notion operation and reason are required.")
        try:
            endpoint = _notion_endpoint(operation, tool_input)
            payload = _notion_payload(operation, tool_input)
            _enforce_operation_size(payload)
        except ValueError as exc:
            return ToolResult(self.name, ActionStatus.FAILED, self.risk_level, str(exc), error=str(exc))
        packet = _api_operation_packet(
            provider="notion",
            operation=operation,
            endpoint=endpoint,
            method=_notion_method(operation),
            payload=payload,
            reason=reason,
            approval_required=operation != "query_database",
            secret_refs={"api_key": "NOTION_API_KEY"},
            metadata={
                "notion_version": str(tool_input.get("notion_version") or "2022-06-28").strip(),
                "target": _notion_target(tool_input),
            },
        )
        path = _api_operation_path(normalized, provider="notion", filename=str(tool_input.get("filename") or ""))
        return _write_api_operation(self.name, self.risk_level, config, path, packet)


class AirtableOperationPrepareTool(Tool):
    def __init__(self) -> None:
        super().__init__(
            name="airtable_operation_prepare",
            description=(
                "Prepare an Airtable API operation artifact for record list/create/update/upsert/delete workflows with schema, "
                "payload preview, approval state, and endpoint metadata. This does not call Airtable."
            ),
            risk_level=RiskLevel.MEDIUM,
            input_schema=object_input_schema(
                {
                    "operation": {"type": "string", "enum": ["list_records", "create_records", "update_records", "upsert_records", "delete_records"]},
                    "base_id": {"type": "string"},
                    "table_name": {"type": "string"},
                    "table_id": {"type": "string"},
                    "records": {"type": "array", "items": {"type": "object"}},
                    "record_ids": {"type": "array", "items": {"type": "string"}},
                    "fields": {"type": "array", "items": {"type": "string"}},
                    "filter_formula": {"type": "string"},
                    "sort": {"type": "array", "items": {"type": "object"}},
                    "upsert_key_fields": {"type": "array", "items": {"type": "string"}},
                    "schema": {"type": "object", "description": "Expected table fields and types used for validation/planning."},
                    "filename": {"type": "string"},
                    "reason": {"type": "string"},
                },
                required=["operation", "base_id", "reason"],
            ),
            capability_group="productivity",
        )

    def execute(self, tool_input: dict[str, Any], config: AgentConfig) -> ToolResult:
        normalized = config.normalized()
        operation = str(tool_input.get("operation") or "").strip()
        base_id = str(tool_input.get("base_id") or "").strip()
        reason = str(tool_input.get("reason") or "").strip()
        table = str(tool_input.get("table_id") or tool_input.get("table_name") or "").strip()
        if not operation or not base_id or not reason:
            return ToolResult(self.name, ActionStatus.FAILED, self.risk_level, "Airtable operation, base_id, and reason are required.")
        if not table:
            return ToolResult(self.name, ActionStatus.FAILED, self.risk_level, "Airtable table_id or table_name is required.")
        try:
            endpoint = _airtable_endpoint(base_id, table)
            payload = _airtable_payload(operation, tool_input)
            _enforce_operation_size(payload)
        except ValueError as exc:
            return ToolResult(self.name, ActionStatus.FAILED, self.risk_level, str(exc), error=str(exc))
        packet = _api_operation_packet(
            provider="airtable",
            operation=operation,
            endpoint=endpoint,
            method=_airtable_method(operation),
            payload=payload,
            reason=reason,
            approval_required=operation != "list_records",
            secret_refs={"api_key": "AIRTABLE_API_KEY"},
            metadata={"base_id": base_id, "table": table, "schema": _json_object(tool_input.get("schema"))},
        )
        path = _api_operation_path(normalized, provider="airtable", filename=str(tool_input.get("filename") or ""))
        return _write_api_operation(self.name, self.risk_level, config, path, packet)


class GoogleWorkspaceOperationPrepareTool(Tool):
    def __init__(self) -> None:
        super().__init__(
            name="google_workspace_operation_prepare",
            description=(
                "Prepare a Google Workspace operation artifact for Calendar, Drive, Docs, Sheets, or Gmail workflows. "
                "This validates the target and payload shape, records OAuth scope needs, and never calls Google APIs."
            ),
            risk_level=RiskLevel.MEDIUM,
            input_schema=object_input_schema(
                {
                    "app": {"type": "string", "enum": ["calendar", "drive", "docs", "sheets", "gmail"]},
                    "operation": {
                        "type": "string",
                        "enum": [
                            "create_event",
                            "update_event",
                            "delete_event",
                            "create_folder",
                            "upload_file",
                            "share_file",
                            "create_doc",
                            "append_doc_text",
                            "create_sheet",
                            "update_values",
                            "append_values",
                            "create_draft",
                        ],
                    },
                    "calendar_id": {"type": "string"},
                    "event_id": {"type": "string"},
                    "file_id": {"type": "string"},
                    "folder_id": {"type": "string"},
                    "document_id": {"type": "string"},
                    "spreadsheet_id": {"type": "string"},
                    "range": {"type": "string"},
                    "title": {"type": "string"},
                    "description": {"type": "string"},
                    "start": {"type": "string"},
                    "end": {"type": "string"},
                    "timezone": {"type": "string"},
                    "attendees": {"type": "array", "items": {"type": "string"}},
                    "recipients": {"type": "array", "items": {"type": "string"}},
                    "role": {"type": "string", "enum": ["reader", "commenter", "writer"]},
                    "path": {"type": "string", "description": "Local file path for upload_file planning."},
                    "mime_type": {"type": "string"},
                    "body": {"type": "string"},
                    "values": {"type": "array", "items": {"type": "array"}},
                    "payload": {"type": "object", "description": "Optional provider-specific payload additions."},
                    "filename": {"type": "string"},
                    "reason": {"type": "string"},
                },
                required=["app", "operation", "reason"],
            ),
            capability_group="productivity",
        )

    def execute(self, tool_input: dict[str, Any], config: AgentConfig) -> ToolResult:
        normalized = config.normalized()
        app = str(tool_input.get("app") or "").strip().lower()
        operation = str(tool_input.get("operation") or "").strip()
        reason = str(tool_input.get("reason") or "").strip()
        if not app or not operation or not reason:
            return ToolResult(self.name, ActionStatus.FAILED, self.risk_level, "Google app, operation, and reason are required.")
        try:
            endpoint = _google_endpoint(app, operation, tool_input)
            payload = _google_payload(app, operation, tool_input)
            _enforce_operation_size(payload)
        except ValueError as exc:
            return ToolResult(self.name, ActionStatus.FAILED, self.risk_level, str(exc), error=str(exc))
        approval_required = _google_approval_required(operation)
        packet = _api_operation_packet(
            provider="google_workspace",
            operation=f"{app}.{operation}",
            endpoint=endpoint,
            method=_google_method(operation),
            payload=payload,
            reason=reason,
            approval_required=approval_required,
            secret_refs={"oauth": "GOOGLE_WORKSPACE_OAUTH_TOKEN"},
            metadata={
                "app": app,
                "operation": operation,
                "scopes": _google_scopes(app, operation),
                "target": _google_target(tool_input),
                "adapter_status": "not_configured_for_live_execution",
            },
        )
        path = _api_operation_path(normalized, provider="google_workspace", filename=str(tool_input.get("filename") or ""))
        return _write_api_operation(self.name, self.risk_level, config, path, packet)


class ApiOperationInspectTool(Tool):
    def __init__(self) -> None:
        super().__init__(
            name="api_operation_inspect",
            description="Inspect a prepared Notion, Airtable, or productivity API operation artifact without executing it.",
            risk_level=RiskLevel.LOW,
            input_schema=object_input_schema({"path": {"type": "string", "description": "Allowed operation artifact JSON path."}}, required=["path"]),
            capability_group="productivity",
        )

    def execute(self, tool_input: dict[str, Any], config: AgentConfig) -> ToolResult:
        normalized = config.normalized()
        path = _resolve_api_operation_path(normalized, str(tool_input.get("path") or ""))
        if not _is_within(path, normalized.allowed_read_roots + normalized.allowed_write_roots):
            return ToolResult(self.name, ActionStatus.BLOCKED, self.risk_level, "API operation path is outside allowed roots.")
        if not path.exists() or path.suffix.lower() != ".json":
            return ToolResult(self.name, ActionStatus.FAILED, self.risk_level, "API operation artifact does not exist.")
        try:
            packet = json.loads(path.read_text(encoding="utf-8"))
        except json.JSONDecodeError as exc:
            return ToolResult(self.name, ActionStatus.FAILED, self.risk_level, "API operation artifact is invalid JSON.", error=str(exc))
        if not isinstance(packet, dict):
            return ToolResult(self.name, ActionStatus.FAILED, self.risk_level, "API operation artifact must be a JSON object.")
        return ToolResult(
            self.name,
            ActionStatus.SUCCEEDED,
            self.risk_level,
            f"Inspected {packet.get('provider', 'api')} operation artifact.",
            {
                "path": str(path),
                "operation_id": packet.get("operation_id", ""),
                "provider": packet.get("provider", ""),
                "operation": packet.get("operation", ""),
                "method": packet.get("method", ""),
                "endpoint": packet.get("endpoint", ""),
                "approval_required": bool(packet.get("approval_required", True)),
                "live_execution_status": packet.get("live_execution_status", ""),
                "payload_shape": _payload_shape(packet.get("payload")),
                "source": "api_operation_inspect",
            },
        )


def default_productivity_tools() -> dict[str, Tool]:
    tools: list[Tool] = [
        EmailDraftPrepareTool(),
        GmailDraftPrepareTool(),
        XlsxWorkbookCreateTool(),
        XlsxWorkbookInspectTool(),
        NotionOperationPrepareTool(),
        AirtableOperationPrepareTool(),
        GoogleWorkspaceOperationPrepareTool(),
        ApiOperationInspectTool(),
    ]
    return {tool.name: tool for tool in tools}


def _email_draft_schema() -> dict[str, Any]:
    return object_input_schema(
        {
            "to": {"type": "array", "items": {"type": "string"}, "description": "Explicit recipient email addresses."},
            "cc": {"type": "array", "items": {"type": "string"}, "description": "Optional CC email addresses."},
            "bcc": {"type": "array", "items": {"type": "string"}, "description": "Optional BCC email addresses."},
            "subject": {"type": "string", "description": "Email subject."},
            "body": {"type": "string", "description": "Email body text."},
            "tone": {"type": "string", "description": "Optional requested tone label."},
            "attachments": {"type": "array", "items": {"type": "string"}, "description": "Optional attachment paths to review before sending."},
            "reason": {"type": "string", "description": "Why this draft should be prepared."},
        },
        required=["to", "subject", "body", "reason"],
    )


def _prepare_email_draft(tool_input: dict[str, Any], config: AgentConfig, *, provider: str) -> ToolResult:
    normalized = config.normalized()
    to = _email_list(tool_input.get("to"))
    cc = _email_list(tool_input.get("cc"))
    bcc = _email_list(tool_input.get("bcc"))
    invalid = [address for address in [*to, *cc, *bcc] if not EMAIL_ADDRESS_RE.match(address)]
    if not to:
        return ToolResult(f"{provider}_draft_prepare", ActionStatus.FAILED, RiskLevel.MEDIUM, "At least one explicit recipient is required.")
    if invalid:
        return ToolResult(f"{provider}_draft_prepare", ActionStatus.FAILED, RiskLevel.MEDIUM, "Draft contains invalid email addresses.", {"invalid_addresses": invalid})
    if len([*to, *cc, *bcc]) > EMAIL_DRAFT_MAX_RECIPIENTS:
        return ToolResult(f"{provider}_draft_prepare", ActionStatus.BLOCKED, RiskLevel.MEDIUM, "Draft exceeds recipient limit.")
    subject = " ".join(str(tool_input.get("subject") or "").split())
    body = str(tool_input.get("body") or "").strip()
    if not subject or not body:
        return ToolResult(f"{provider}_draft_prepare", ActionStatus.FAILED, RiskLevel.MEDIUM, "Subject and body are required.")
    if len(body) > EMAIL_DRAFT_MAX_BODY_CHARS:
        return ToolResult(f"{provider}_draft_prepare", ActionStatus.BLOCKED, RiskLevel.MEDIUM, "Draft body exceeds safety limit.")
    draft_id = f"{provider}-draft-{datetime.now(timezone.utc).strftime('%Y%m%d-%H%M%S')}-{uuid4().hex[:8]}"
    root = normalized.data_dir / "email_drafts"
    json_path = (root / f"{draft_id}.json").resolve()
    body_path = (root / f"{draft_id}.txt").resolve()
    if not _is_within(json_path, normalized.allowed_write_roots):
        return ToolResult(f"{provider}_draft_prepare", ActionStatus.BLOCKED, RiskLevel.MEDIUM, "Draft path is outside allowed write roots.")
    attachments = [str(item).strip() for item in (tool_input.get("attachments") or []) if str(item).strip()]
    draft = {
        "draft_id": draft_id,
        "provider": provider,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "to": to,
        "cc": cc,
        "bcc": bcc,
        "subject": subject,
        "body": body,
        "tone": str(tool_input.get("tone") or "").strip(),
        "attachments": attachments,
        "reason": str(tool_input.get("reason") or "").strip(),
        "send_status": "not_sent",
        "approval_required_for_send": True,
    }
    draft["gmail_command_preview"] = [
        "gog",
        "gmail",
        "drafts",
        "create",
        "--to",
        ",".join(to),
        "--subject",
        subject,
        "--body-file",
        str(body_path),
    ] if provider == "gmail" else []
    if config.dry_run:
        return ToolResult(
            f"{provider}_draft_prepare",
            ActionStatus.SKIPPED,
            RiskLevel.MEDIUM,
            f"Dry run: would prepare {provider} draft.",
            {"draft": draft, "path": str(json_path), "body_path": str(body_path)},
        )
    root.mkdir(parents=True, exist_ok=True)
    json_path.write_text(json.dumps(draft, indent=2, ensure_ascii=False, sort_keys=True) + "\n", encoding="utf-8")
    body_path.write_text(body + "\n", encoding="utf-8")
    return ToolResult(
        f"{provider}_draft_prepare",
        ActionStatus.SUCCEEDED,
        RiskLevel.MEDIUM,
        f"Prepared {provider} draft artifact.",
        {"draft": {**draft, "body": body[:500], "body_truncated": len(body) > 500}, "path": str(json_path), "body_path": str(body_path)},
    )


def _email_list(value: Any) -> list[str]:
    if isinstance(value, str):
        raw = [value]
    elif isinstance(value, list):
        raw = value
    else:
        raw = []
    cleaned = []
    seen = set()
    for item in raw:
        address = str(item).strip().lower()
        if address and address not in seen:
            seen.add(address)
            cleaned.append(address)
    return cleaned


def _safe_filename(value: str, suffix: str) -> str:
    name = Path(value).name.strip() or f"artifact{suffix}"
    if not name.lower().endswith(suffix):
        name += suffix
    stem = "".join(char if char.isalnum() or char in ("-", "_", ".") else "-" for char in Path(name).stem).strip(".-")
    return f"{stem or 'artifact'}{suffix}"


def _safe_sheet_name(value: str, existing: list[str]) -> str:
    cleaned = "".join(char for char in value if char not in r"[]:*?/\\").strip()[:31] or "Sheet"
    candidate = cleaned
    index = 2
    while candidate in existing:
        suffix = f" {index}"
        candidate = f"{cleaned[:31 - len(suffix)]}{suffix}"
        index += 1
    return candidate


def _resolve_allowed_path(config: AgentConfig, raw_path: str) -> Path:
    path = Path(raw_path).expanduser()
    if not path.is_absolute():
        path = config.workspace / path
        if not path.exists():
            data_path = config.data_dir / raw_path
            if data_path.exists():
                path = data_path
            else:
                spreadsheet_path = config.data_dir / "spreadsheets" / Path(raw_path).name
                if spreadsheet_path.exists():
                    path = spreadsheet_path
    return path.resolve()


def _xlsx_cell_value(value: Any) -> Any:
    if isinstance(value, dict):
        if "formula" in value:
            formula = str(value.get("formula") or "").strip()
            if formula:
                return formula if formula.startswith("=") else f"={formula}"
            return None
        if "value" in value:
            return value.get("value")
    return value


def _notion_endpoint(operation: str, tool_input: dict[str, Any]) -> str:
    if operation == "create_page":
        if not str(tool_input.get("database_id") or tool_input.get("parent_page_id") or "").strip():
            raise ValueError("create_page requires database_id or parent_page_id.")
        return "https://api.notion.com/v1/pages"
    if operation == "update_page":
        page_id = str(tool_input.get("page_id") or "").strip()
        if not page_id:
            raise ValueError("update_page requires page_id.")
        return f"https://api.notion.com/v1/pages/{page_id}"
    if operation == "append_blocks":
        page_id = str(tool_input.get("page_id") or "").strip()
        if not page_id:
            raise ValueError("append_blocks requires page_id.")
        return f"https://api.notion.com/v1/blocks/{page_id}/children"
    if operation == "query_database":
        database_id = str(tool_input.get("database_id") or "").strip()
        if not database_id:
            raise ValueError("query_database requires database_id.")
        return f"https://api.notion.com/v1/databases/{database_id}/query"
    if operation == "update_database_schema":
        database_id = str(tool_input.get("database_id") or "").strip()
        if not database_id:
            raise ValueError("update_database_schema requires database_id.")
        return f"https://api.notion.com/v1/databases/{database_id}"
    raise ValueError(f"Unsupported Notion operation: {operation}")


def _notion_method(operation: str) -> str:
    if operation in {"update_page", "update_database_schema"}:
        return "PATCH"
    return "POST"


def _notion_payload(operation: str, tool_input: dict[str, Any]) -> dict[str, Any]:
    properties = _json_object(tool_input.get("properties"))
    blocks = _json_list(tool_input.get("blocks"), limit=API_OPERATION_MAX_BLOCKS)
    if operation == "create_page":
        parent = {"database_id": str(tool_input.get("database_id") or "").strip()} if tool_input.get("database_id") else {"page_id": str(tool_input.get("parent_page_id") or "").strip()}
        payload = {"parent": parent, "properties": properties}
        if tool_input.get("title") and "title" not in payload["properties"]:
            payload["properties"]["title"] = {"title": [{"text": {"content": str(tool_input.get("title")).strip()}}]}
        if blocks:
            payload["children"] = blocks
        return payload
    if operation == "update_page":
        if not properties:
            raise ValueError("update_page requires properties.")
        return {"properties": properties}
    if operation == "append_blocks":
        if not blocks:
            raise ValueError("append_blocks requires blocks.")
        return {"children": blocks}
    if operation == "query_database":
        payload: dict[str, Any] = {}
        if isinstance(tool_input.get("filter"), dict):
            payload["filter"] = tool_input["filter"]
        sorts = _json_list(tool_input.get("sorts"), limit=20)
        if sorts:
            payload["sorts"] = sorts
        return payload
    if operation == "update_database_schema":
        schema = _json_object(tool_input.get("schema"))
        if not schema:
            raise ValueError("update_database_schema requires schema.")
        return {"properties": schema}
    raise ValueError(f"Unsupported Notion operation: {operation}")


def _notion_target(tool_input: dict[str, Any]) -> dict[str, str]:
    return {
        "page_id": str(tool_input.get("page_id") or "").strip(),
        "database_id": str(tool_input.get("database_id") or "").strip(),
        "parent_page_id": str(tool_input.get("parent_page_id") or "").strip(),
    }


def _airtable_endpoint(base_id: str, table: str) -> str:
    return f"https://api.airtable.com/v0/{base_id}/{table}"


def _airtable_method(operation: str) -> str:
    if operation == "list_records":
        return "GET"
    if operation in {"update_records", "upsert_records"}:
        return "PATCH"
    if operation == "delete_records":
        return "DELETE"
    return "POST"


def _airtable_payload(operation: str, tool_input: dict[str, Any]) -> dict[str, Any]:
    records = _json_list(tool_input.get("records"), limit=API_OPERATION_MAX_RECORDS)
    record_ids = _string_list(tool_input.get("record_ids"), limit=API_OPERATION_MAX_RECORDS)
    if operation == "list_records":
        payload: dict[str, Any] = {}
        fields = _string_list(tool_input.get("fields"), limit=100)
        if fields:
            payload["fields"] = fields
        if tool_input.get("filter_formula"):
            payload["filterByFormula"] = str(tool_input.get("filter_formula")).strip()
        sort = _json_list(tool_input.get("sort"), limit=20)
        if sort:
            payload["sort"] = sort
        return payload
    if operation == "create_records":
        if not records:
            raise ValueError("create_records requires records.")
        return {"records": [_airtable_record(record, require_id=False) for record in records]}
    if operation == "update_records":
        if not records:
            raise ValueError("update_records requires records with id and fields.")
        return {"records": [_airtable_record(record, require_id=True) for record in records]}
    if operation == "upsert_records":
        key_fields = _string_list(tool_input.get("upsert_key_fields"), limit=20)
        if not records or not key_fields:
            raise ValueError("upsert_records requires records and upsert_key_fields.")
        return {"performUpsert": {"fieldsToMergeOn": key_fields}, "records": [_airtable_record(record, require_id=False) for record in records]}
    if operation == "delete_records":
        if not record_ids:
            raise ValueError("delete_records requires record_ids.")
        return {"records": record_ids}
    raise ValueError(f"Unsupported Airtable operation: {operation}")


def _airtable_record(record: Any, *, require_id: bool) -> dict[str, Any]:
    if not isinstance(record, dict):
        raise ValueError("Airtable records must be objects.")
    record_id = str(record.get("id") or "").strip()
    fields = record.get("fields")
    if fields is None:
        fields = {key: value for key, value in record.items() if key != "id"}
    if require_id and not record_id:
        raise ValueError("Airtable update records require id.")
    if not isinstance(fields, dict) or not fields:
        raise ValueError("Airtable records require fields.")
    payload = {"fields": fields}
    if record_id:
        payload["id"] = record_id
    return payload


def _google_endpoint(app: str, operation: str, tool_input: dict[str, Any]) -> str:
    if app == "calendar":
        calendar_id = str(tool_input.get("calendar_id") or "primary").strip()
        if operation == "create_event":
            return f"https://www.googleapis.com/calendar/v3/calendars/{calendar_id}/events"
        if operation in {"update_event", "delete_event"}:
            event_id = str(tool_input.get("event_id") or "").strip()
            if not event_id:
                raise ValueError(f"{operation} requires event_id.")
            return f"https://www.googleapis.com/calendar/v3/calendars/{calendar_id}/events/{event_id}"
    if app == "drive":
        if operation == "create_folder":
            return "https://www.googleapis.com/drive/v3/files"
        if operation == "upload_file":
            return "https://www.googleapis.com/upload/drive/v3/files?uploadType=multipart"
        if operation == "share_file":
            file_id = str(tool_input.get("file_id") or "").strip()
            if not file_id:
                raise ValueError("share_file requires file_id.")
            return f"https://www.googleapis.com/drive/v3/files/{file_id}/permissions"
    if app == "docs":
        if operation == "create_doc":
            return "https://docs.googleapis.com/v1/documents"
        if operation == "append_doc_text":
            document_id = str(tool_input.get("document_id") or "").strip()
            if not document_id:
                raise ValueError("append_doc_text requires document_id.")
            return f"https://docs.googleapis.com/v1/documents/{document_id}:batchUpdate"
    if app == "sheets":
        if operation == "create_sheet":
            return "https://sheets.googleapis.com/v4/spreadsheets"
        if operation in {"update_values", "append_values"}:
            spreadsheet_id = str(tool_input.get("spreadsheet_id") or "").strip()
            value_range = str(tool_input.get("range") or "").strip()
            if not spreadsheet_id or not value_range:
                raise ValueError(f"{operation} requires spreadsheet_id and range.")
            action = "append" if operation == "append_values" else "values"
            suffix = f":append" if operation == "append_values" else ""
            return f"https://sheets.googleapis.com/v4/spreadsheets/{spreadsheet_id}/values/{value_range}{suffix}"
    if app == "gmail" and operation == "create_draft":
        return "https://gmail.googleapis.com/gmail/v1/users/me/drafts"
    raise ValueError(f"Unsupported Google Workspace operation: {app}.{operation}")


def _google_method(operation: str) -> str:
    if operation in {"update_event", "update_values"}:
        return "PUT"
    if operation == "delete_event":
        return "DELETE"
    return "POST"


def _google_payload(app: str, operation: str, tool_input: dict[str, Any]) -> dict[str, Any]:
    extra = _json_object(tool_input.get("payload"))
    if app == "calendar":
        if operation == "delete_event":
            return extra
        title = str(tool_input.get("title") or "").strip()
        start = str(tool_input.get("start") or "").strip()
        end = str(tool_input.get("end") or "").strip()
        if not title or not start or not end:
            raise ValueError(f"{operation} requires title, start, and end.")
        payload = {
            "summary": title,
            "description": str(tool_input.get("description") or "").strip(),
            "start": {"dateTime": start, "timeZone": str(tool_input.get("timezone") or "UTC").strip()},
            "end": {"dateTime": end, "timeZone": str(tool_input.get("timezone") or "UTC").strip()},
        }
        attendees = _email_list(tool_input.get("attendees"))
        if attendees:
            payload["attendees"] = [{"email": address} for address in attendees]
        payload.update(extra)
        return payload
    if app == "drive":
        title = str(tool_input.get("title") or "").strip()
        if operation == "create_folder":
            if not title:
                raise ValueError("create_folder requires title.")
            payload = {"name": title, "mimeType": "application/vnd.google-apps.folder"}
            folder_id = str(tool_input.get("folder_id") or "").strip()
            if folder_id:
                payload["parents"] = [folder_id]
            payload.update(extra)
            return payload
        if operation == "upload_file":
            local_path = str(tool_input.get("path") or "").strip()
            if not title or not local_path:
                raise ValueError("upload_file requires title and path.")
            payload = {
                "metadata": {
                    "name": title,
                    "mimeType": str(tool_input.get("mime_type") or "application/octet-stream").strip(),
                    "parents": [str(tool_input.get("folder_id")).strip()] if tool_input.get("folder_id") else [],
                },
                "local_path": local_path,
                "upload_status": "not_uploaded",
            }
            payload.update(extra)
            return payload
        if operation == "share_file":
            recipients = _email_list(tool_input.get("recipients"))
            if not recipients:
                raise ValueError("share_file requires recipients.")
            role = str(tool_input.get("role") or "reader").strip()
            return {"permissions": [{"type": "user", "role": role, "emailAddress": address} for address in recipients], **extra}
    if app == "docs":
        title = str(tool_input.get("title") or "").strip()
        body = str(tool_input.get("body") or "").strip()
        if operation == "create_doc":
            if not title:
                raise ValueError("create_doc requires title.")
            payload = {"title": title}
            if body:
                payload["initial_text"] = body
                payload["follow_up_batch_update_required"] = True
            payload.update(extra)
            return payload
        if operation == "append_doc_text":
            if not body:
                raise ValueError("append_doc_text requires body.")
            payload = {"requests": [{"insertText": {"location": {"index": 1}, "text": body}}]}
            payload.update(extra)
            return payload
    if app == "sheets":
        title = str(tool_input.get("title") or "").strip()
        values = _json_list(tool_input.get("values"), limit=API_OPERATION_MAX_RECORDS)
        if operation == "create_sheet":
            if not title:
                raise ValueError("create_sheet requires title.")
            payload = {"properties": {"title": title}}
            if values:
                payload["initial_values"] = values
            payload.update(extra)
            return payload
        if operation in {"update_values", "append_values"}:
            if not values:
                raise ValueError(f"{operation} requires values.")
            payload = {"range": str(tool_input.get("range") or "").strip(), "majorDimension": "ROWS", "values": values}
            payload.update(extra)
            return payload
    if app == "gmail" and operation == "create_draft":
        recipients = _email_list(tool_input.get("recipients") or tool_input.get("to"))
        title = str(tool_input.get("title") or tool_input.get("subject") or "").strip()
        body = str(tool_input.get("body") or "").strip()
        if not recipients or not title or not body:
            raise ValueError("create_draft requires recipients, title/subject, and body.")
        return {"message": {"to": recipients, "subject": title, "body": body, "raw_status": "not_encoded"}, **extra}
    raise ValueError(f"Unsupported Google Workspace payload: {app}.{operation}")


def _google_scopes(app: str, operation: str) -> list[str]:
    if app == "calendar":
        return ["https://www.googleapis.com/auth/calendar.events"]
    if app == "drive":
        return ["https://www.googleapis.com/auth/drive.file"]
    if app == "docs":
        return ["https://www.googleapis.com/auth/documents"]
    if app == "sheets":
        return ["https://www.googleapis.com/auth/spreadsheets"]
    if app == "gmail":
        return ["https://www.googleapis.com/auth/gmail.compose"]
    return []


def _google_approval_required(operation: str) -> bool:
    return operation not in set()


def _google_target(tool_input: dict[str, Any]) -> dict[str, str]:
    return {
        "calendar_id": str(tool_input.get("calendar_id") or "").strip(),
        "event_id": str(tool_input.get("event_id") or "").strip(),
        "file_id": str(tool_input.get("file_id") or "").strip(),
        "folder_id": str(tool_input.get("folder_id") or "").strip(),
        "document_id": str(tool_input.get("document_id") or "").strip(),
        "spreadsheet_id": str(tool_input.get("spreadsheet_id") or "").strip(),
        "range": str(tool_input.get("range") or "").strip(),
    }


def _api_operation_packet(
    *,
    provider: str,
    operation: str,
    endpoint: str,
    method: str,
    payload: dict[str, Any],
    reason: str,
    approval_required: bool,
    secret_refs: dict[str, str],
    metadata: dict[str, Any],
) -> dict[str, Any]:
    return {
        "operation_id": f"{provider}-operation-{datetime.now(timezone.utc).strftime('%Y%m%d-%H%M%S')}-{uuid4().hex[:8]}",
        "created_at": datetime.now(timezone.utc).isoformat(),
        "provider": provider,
        "operation": operation,
        "method": method,
        "endpoint": endpoint,
        "payload": payload,
        "metadata": metadata,
        "reason": reason,
        "secret_refs": secret_refs,
        "approval_required": approval_required,
        "live_execution_status": "not_executed",
        "safety_note": "Prepared artifact only. No external API call was made and no remote data was mutated.",
    }


def _api_operation_path(config: AgentConfig, *, provider: str, filename: str) -> Path:
    safe = _safe_filename(filename or f"{provider}-operation-{uuid4().hex[:8]}.json", ".json")
    return (config.data_dir / "api_operations" / provider / safe).resolve()


def _write_api_operation(tool_name: str, risk_level: RiskLevel, config: AgentConfig, path: Path, packet: dict[str, Any]) -> ToolResult:
    normalized = config.normalized()
    if not _is_within(path, normalized.allowed_write_roots):
        return ToolResult(tool_name, ActionStatus.BLOCKED, risk_level, "API operation path is outside allowed write roots.")
    if config.dry_run:
        return ToolResult(tool_name, ActionStatus.SKIPPED, risk_level, f"Dry run: would prepare {packet['provider']} operation.", {"path": str(path), "operation": packet})
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(packet, indent=2, ensure_ascii=False, sort_keys=True) + "\n", encoding="utf-8")
    return ToolResult(
        tool_name,
        ActionStatus.SUCCEEDED,
        risk_level,
        f"Prepared {packet['provider']} {packet['operation']} operation artifact.",
        {
            "path": str(path),
            "operation_id": packet["operation_id"],
            "provider": packet["provider"],
            "operation": packet["operation"],
            "method": packet["method"],
            "endpoint": packet["endpoint"],
            "approval_required": packet["approval_required"],
            "live_execution_status": packet["live_execution_status"],
            "source": f"{packet['provider']}_operation_prepare",
        },
    )


def _resolve_api_operation_path(config: AgentConfig, raw_path: str) -> Path:
    path = Path(raw_path).expanduser()
    if not path.is_absolute():
        path = config.workspace / path
        if not path.exists():
            data_path = config.data_dir / raw_path
            if data_path.exists():
                path = data_path
            else:
                path = config.data_dir / "api_operations" / Path(raw_path).name
    return path.resolve()


def _enforce_operation_size(payload: dict[str, Any]) -> None:
    encoded = json.dumps(payload, ensure_ascii=False, sort_keys=True)
    if len(encoded) > API_OPERATION_BODY_CHARS:
        raise ValueError("Prepared API payload exceeds safety limit.")


def _json_object(value: Any) -> dict[str, Any]:
    return dict(value) if isinstance(value, dict) else {}


def _json_list(value: Any, *, limit: int) -> list[Any]:
    if not isinstance(value, list):
        return []
    return value[: max(0, limit)]


def _string_list(value: Any, *, limit: int) -> list[str]:
    if isinstance(value, str):
        return [value.strip()] if value.strip() else []
    if not isinstance(value, list):
        return []
    return [str(item).strip() for item in value[:limit] if str(item).strip()]


def _payload_shape(payload: Any) -> dict[str, Any]:
    if isinstance(payload, dict):
        return {"type": "object", "keys": sorted(payload.keys())[:30]}
    if isinstance(payload, list):
        return {"type": "array", "length": len(payload)}
    return {"type": type(payload).__name__}


def _is_within(path: Path, roots: tuple[Path, ...]) -> bool:
    return any(path == root or root in path.parents for root in roots)
